# 管理用户账户信息
from logging import error
from tokenize import String, group
import pandas as pd
from openpyxl import load_workbook
import os
import datetime
import numpy as np

cur_user = ''

# 用户名密码是否输入正确
def isCorrect(username: String, userpassword: String):
    if not os.path.exists("./userInfo.xlsx"):
        return False
    else:
        df = pd.DataFrame(pd.read_excel('./userInfo.xlsx', sheet_name="userInfo"))
        result = df[(df.username==username) & (df.password==userpassword)]
        return not result.empty

# 该用户名是否已被注册
def isValid(username: String):
    if not os.path.exists("./userInfo.xlsx"):
        return True
    else:
        df = pd.DataFrame(pd.read_excel('./userInfo.xlsx', sheet_name="userInfo"))
        userList = df['username']
        if username not in userList.values:
            return True
    return False


# 注册新用户
def createNewUser(username: String, userpassword: String):
    if not os.path.exists("./userInfo.xlsx"):
        df = pd.DataFrame([{'username' : username, 'password' : userpassword}])
        df.to_excel("./userInfo.xlsx", sheet_name="userInfo", index=False)
        return
    else:
        df_old = pd.DataFrame(pd.read_excel('./userInfo.xlsx', sheet_name="userInfo"))
        n = len(df_old.index)
        df = pd.DataFrame([{'username' : username, 'password' : userpassword}])
        book = load_workbook('./userInfo.xlsx')
        group_writer = pd.ExcelWriter('./userInfo.xlsx', engine='openpyxl')
        group_writer.book = book
        group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_rows = df_old.shape[0]
        df.to_excel(group_writer, sheet_name='userInfo', startrow=df_rows+1,header=False, index=False)
        group_writer.save()
        return

def saveLanguage(src, language):
    if not os.path.exists("./userInfo.xlsx"):
        raise error
    username, proname = src.split('/')
    book = load_workbook('./userInfo.xlsx')
    group_writer = pd.ExcelWriter('./userInfo.xlsx', engine='openpyxl')
    group_writer.book = book
    group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    t = datetime.datetime.now()
    if username in group_writer.sheets:
        group_writer.close()
        df_old = pd.DataFrame(pd.read_excel('./userInfo.xlsx', sheet_name=username))
        df = pd.DataFrame([{'proname': proname, 'language': language, 'lastupdate': t}])
        book = load_workbook('./userInfo.xlsx')
        group_writer = pd.ExcelWriter('./userInfo.xlsx', engine='openpyxl')
        group_writer.book = book
        group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_rows = df_old.shape[0]
        df.to_excel(group_writer, sheet_name=username, startrow=df_rows+1,header=False, index=False)
        group_writer.save()
    else:
        df = pd.DataFrame([{'proname': proname, 'language': language, 'lastupdate': t}])
        df.to_excel(group_writer, sheet_name=username, index=False)
        group_writer.save()
    return t

def deletepro(src):
    if not os.path.exists("./userInfo.xlsx"):
        raise error
    username, proname = src.split('/')
    df = pd.DataFrame(pd.read_excel('./userInfo.xlsx', sheet_name=username))
    df_del = df[(df.proname!=proname)]
    book = load_workbook('./userInfo.xlsx')
    with pd.ExcelWriter('./userInfo.xlsx', engine='openpyxl') as writer:
        writer.book = book
        idx = book.sheetnames.index(username)
        book.remove(book.worksheets[idx])
        book.create_sheet(username, idx)
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_del.to_excel(writer, sheet_name=username, index=False)

def renamepro(src, dst):
    if not os.path.exists("./userInfo.xlsx"):
        raise error
    username, proname = src.split('/')
    _, newname = dst.split('/')
    df = pd.DataFrame(pd.read_excel('userInfo.xlsx', sheet_name=username))
    df_del = df[(df.proname!=proname)]
    df_old = df[(df.proname==proname)].squeeze()
    tmp = pd.DataFrame([{'proname': newname, 'language': df_old['language'], 'lastupdate': datetime.datetime.now()}])
    frame = [df_del, tmp]
    result = pd.concat(frame)
    book = load_workbook('./userInfo.xlsx')
    with pd.ExcelWriter('./userInfo.xlsx', engine='openpyxl') as writer:
        writer.book = book
        idx = book.sheetnames.index(username)
        book.remove(book.worksheets[idx])
        book.create_sheet(username, idx)
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        result.to_excel(writer, sheet_name=username, index=False)

def showpro(username):
    if not os.path.exists("./userInfo.xlsx"):
        raise error
    try:
        df = pd.DataFrame(pd.read_excel('userInfo.xlsx', sheet_name=username))
        ls = df.values.tolist()
    except:
        ls = []
    return ls

def setCurUser(username: String):
    global cur_user
    cur_user = username

def currentUser():
    return cur_user